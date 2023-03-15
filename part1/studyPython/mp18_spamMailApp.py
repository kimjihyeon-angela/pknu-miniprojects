# 대량 메일 전송
# pip instal openpyxl
from openpyxl import load_workbook 
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

wbook = load_workbook('./studyPython/SpamMailList.xlsx', data_only=True)
wsheet = wbook.active  # sheet1 선택됨

for i in range(1, wsheet.max_row+1):
    recv_mail = wsheet.cell(i, 1).value
    print(recv_mail)
    try:
        # 실제 메일 전송 로직
        send_mail = 'angela9830@naver.com'
        send_pass = ''
        smtp_name = 'smtp.naver.com'
        smtp_port = 587
        msg = MIMEMultipart()
        msg['Subject'] = '엑셀에서 보내는 메일!!'
        msg['From'] = send_mail
        msg['To'] = recv_mail
        text = '''보내는 내용입니다. 메롱~!!!'''
        msg.attach(MIMEText(text))

        mail = smtplib.SMTP(smtp_name, smtp_port)
        mail.starttls() # 보안
        mail.login(send_mail, send_pass)
        mail.sendmail(send_mail, recv_mail, msg.as_string())
        mail.quit()
        print(f'전송 성공 : {recv_mail}')
    except Exception as e:
        print(f'수신메일 - {recv_mail}')
        print(f'전송에러 : {e}')
